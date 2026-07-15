"""Generate an Excel profit report from the item database.

One row per listed item: cover photo, title, our price, shipping, Ross cost, and
a Net profit column that nets out eBay's final value fee, the Promoted Listings
ad rate, and shipping. The fee/shipping assumptions are editable cells at the top
and the per-row math is written as Excel FORMULAS, so changing an assumption
recalculates every row live in Excel.

Run standalone (`python report.py`) to write report.xlsx, or use the bot's
/report command, which builds this and sends it to you on Telegram.
"""
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from db import list_items

try:
    from PIL import Image, ImageOps
except ImportError:
    Image = None
    ImageOps = None

# Items with a real listing price — everything from drafting onward.
_REPORTABLE = {"drafted", "review", "approved", "ebay_draft", "published", "sold"}

MONEY = "$#,##0.00"
PCT = "0.00%"
THUMB_PX = 92

# Assumption cell addresses (column C, rows 4-8) — referenced by the row formulas.
A_FVF, A_FIXED, A_AD, A_SHIP_CHG, A_SHIP_COST = "$C$4", "$C$5", "$C$6", "$C$7", "$C$8"


def _price_for(item):
    """Sale price if the item sold, else its listed price."""
    ebay = item.get("ebay") or {}
    if ebay.get("sale_price") is not None:
        return ebay["sale_price"]
    return (item.get("listing") or {}).get("price")


def _thumbnail(cover_path, dest_dir):
    if Image is None or not cover_path:
        return None
    p = Path(cover_path)
    if not p.exists():
        return None
    try:
        img = ImageOps.exif_transpose(Image.open(p)).convert("RGB")
        img.thumbnail((THUMB_PX, THUMB_PX))
        out = Path(dest_dir) / (p.stem + ".png")
        img.save(out, "PNG")
        return str(out)
    except Exception:
        return None


def build_report(path: str) -> str:
    items = [i for i in list_items()
             if (i.get("listing") or {}).get("price") is not None and i["status"] in _REPORTABLE]
    items.sort(key=lambda i: (i["status"], i.get("created_at", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "Profit"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    edit_fill = PatternFill("solid", fgColor="FFF2CC")   # editable assumption cells
    missing_fill = PatternFill("solid", fgColor="FCE4D6")  # unknown cost — fill it in
    total_fill = PatternFill("solid", fgColor="E2EFDA")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Ross Resale — Profit Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Generated " + datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
                + "  ·  edit the yellow assumption cells to recalc every row")

    # --- Assumptions (editable, rows 4-8) ---
    assumptions = [
        ("eBay final value fee %", 0.1325, PCT),
        ("eBay fixed fee per order", 0.40, MONEY),
        ("Promoted Listings ad rate %", 0.04, PCT),
        ("Shipping charged to buyer", 10.0, MONEY),
        ("Your shipping cost (postage)", 10.0, MONEY),
    ]
    for idx, (label, val, fmt) in enumerate(assumptions):
        r = 4 + idx
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = bold
        cell = ws[f"C{r}"]
        cell.value = val
        cell.number_format = fmt
        cell.fill = edit_fill
        cell.border = border

    # --- Table header (row 10) ---
    headers = ["Photo", "ID", "Title", "Status", "Price", "Ship chg", "eBay fee",
               "Ad fee", "Ship cost", "Cost (Ross)", "Net profit", "Margin"]
    HR = 10
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=HR, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    tmpdir = tempfile.mkdtemp(prefix="report_thumbs_")
    first = HR + 1
    r = first
    for item in items:
        L = item.get("listing") or {}
        price = round(float(_price_for(item)), 2)
        paid = (item.get("receipt") or {}).get("reduced_price")

        ws.row_dimensions[r].height = 72
        photos = item.get("photos") or []
        thumb = _thumbnail(photos[0] if photos else None, tmpdir)
        if thumb:
            ws.add_image(XLImage(thumb), f"A{r}")

        ws.cell(row=r, column=2, value=item["item_id"][:8])
        ws.cell(row=r, column=3, value=L.get("title") or "(no title)")
        ws.cell(row=r, column=4, value=item["status"])
        ws.cell(row=r, column=5, value=price).number_format = MONEY
        ws.cell(row=r, column=6, value=f"={A_SHIP_CHG}").number_format = MONEY
        ws.cell(row=r, column=7, value=f"={A_FVF}*(E{r}+F{r})+{A_FIXED}").number_format = MONEY
        ws.cell(row=r, column=8, value=f"={A_AD}*E{r}").number_format = MONEY
        ws.cell(row=r, column=9, value=f"={A_SHIP_COST}").number_format = MONEY
        cost_cell = ws.cell(row=r, column=10)
        cost_cell.number_format = MONEY
        if paid is not None:
            cost_cell.value = round(float(paid), 2)
        else:
            cost_cell.fill = missing_fill  # unknown — treated as 0 until you fill it
        ws.cell(row=r, column=11, value=f"=E{r}+F{r}-G{r}-H{r}-I{r}-J{r}").number_format = MONEY
        ws.cell(row=r, column=12, value=f'=IF((E{r}+F{r})=0,"",K{r}/(E{r}+F{r}))').number_format = "0.0%"
        for col in range(2, 13):
            ws.cell(row=r, column=col).border = border
        r += 1

    # --- Totals row ---
    last = r - 1
    if last >= first:
        ws.cell(row=r, column=2, value="TOTAL").font = bold
        for col_letter in ("E", "G", "H", "I", "J", "K"):
            c = ws.cell(row=r, column=ord(col_letter) - 64,
                        value=f"=SUM({col_letter}{first}:{col_letter}{last})")
            c.number_format = MONEY
            c.font = bold
        for col in range(2, 13):
            cell = ws.cell(row=r, column=col)
            cell.fill = total_fill
            cell.border = border

    # --- Widths + note ---
    widths = {"A": 14, "B": 10, "C": 46, "D": 11, "E": 10, "F": 10, "G": 10,
              "H": 9, "I": 10, "J": 12, "K": 12, "L": 9}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A11"

    note_row = r + 2
    ws.cell(row=note_row, column=2,
            value="Orange 'Cost (Ross)' cells = no receipt was captured; type the amount you paid to complete the profit.")
    ws.cell(row=note_row, column=2).font = Font(italic=True, color="808080")

    wb.save(path)
    return path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "report.xlsx"
    build_report(out)
    print(f"Wrote {out}")
